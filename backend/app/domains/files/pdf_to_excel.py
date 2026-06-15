"""PDF to Excel MVP conversion helpers."""

from __future__ import annotations

import os
import re
from pathlib import Path

from fastapi import HTTPException, status
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.domains.files.pdf_text import TextPDFError, extract_text_pdf_pages


PDF_TO_EXCEL_MIN_TEXT_CHARS = 24
PDF_TO_EXCEL_MAX_PAGES = 60


class PDFToExcelError(TextPDFError):
    """User-facing conversion error for PDF to Excel Beta."""


def convert_text_pdf_to_xlsx(*, input_path: str, output_path: str) -> dict:
    """Extract simple row/column text from a text-based PDF and write XLSX."""

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    try:
        source, page_texts, total_chars = extract_text_pdf_pages(
            input_path=input_path,
            min_text_chars=PDF_TO_EXCEL_MIN_TEXT_CHARS,
            max_pages=PDF_TO_EXCEL_MAX_PAGES,
            product_label="PDF to Excel Beta",
        )
    except TextPDFError as exc:
        raise PDFToExcelError(str(exc)) from exc

    workbook = Workbook()
    summary = workbook.active
    summary.title = "Extracted text"
    _write_header(summary, ["Page", "Row", "Column A", "Column B", "Column C", "Column D", "Column E", "Column F"])

    row_index = 2
    max_columns = 1
    data_rows = 0
    for page_number, text in enumerate(page_texts, start=1):
        for logical_row in _extract_rows(text):
            cells = _split_row_into_cells(logical_row)
            max_columns = max(max_columns, len(cells))
            summary.cell(row=row_index, column=1, value=page_number)
            summary.cell(row=row_index, column=2, value=data_rows + 1)
            for cell_index, value in enumerate(cells[:6], start=3):
                summary.cell(row=row_index, column=cell_index, value=value)
            row_index += 1
            data_rows += 1

    if data_rows == 0:
        raise PDFToExcelError(
            "No table-like text was found. PDF to Excel Beta works best with simple text tables."
        )

    _autosize_columns(summary)
    meta = workbook.create_sheet("Notes")
    meta["A1"] = "PDF to Excel Beta"
    meta["A1"].font = Font(bold=True)
    meta["A2"] = "Source"
    meta["B2"] = source.name
    meta["A3"] = "Pages"
    meta["B3"] = len(page_texts)
    meta["A4"] = "Rows"
    meta["B4"] = data_rows
    meta["A5"] = "Limit"
    meta["B5"] = "Simple text tables only; complex layout, merged cells, and scans are not guaranteed."
    _autosize_columns(meta)

    workbook.save(str(output))
    return {
        "success": True,
        "output_path": str(output),
        "file_size": os.path.getsize(output),
        "page_count": len(page_texts),
        "row_count": data_rows,
        "max_detected_columns": max_columns,
        "extracted_characters": total_chars,
    }


def validate_pdf_to_excel_upload(file_id: str, filename: str | None) -> None:
    if not file_id:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="file_id is required")
    if filename and not filename.lower().endswith(".pdf"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="PDF to Excel Beta only accepts PDF files.",
        )


def user_facing_pdf_to_excel_error(exc: Exception) -> str:
    if isinstance(exc, PDFToExcelError):
        return str(exc)
    message = str(exc).splitlines()[0].strip()
    return message[:240] if message else "PDF to Excel conversion failed."


def _extract_rows(text: str) -> list[str]:
    rows: list[str] = []
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        rows.append(line)
    return rows


def _split_row_into_cells(row: str) -> list[str]:
    if "\t" in row:
        cells = row.split("\t")
    elif re.search(r"\s{2,}", row):
        cells = re.split(r"\s{2,}", row)
    elif "," in row and row.count(",") <= 12:
        cells = row.split(",")
    elif "|" in row:
        cells = row.split("|")
    else:
        cells = [row]
    return [cell.strip() for cell in cells if cell.strip()]


def _write_header(sheet, values: list[str]) -> None:
    fill = PatternFill("solid", fgColor="E0F2FE")
    for column, value in enumerate(values, start=1):
        cell = sheet.cell(row=1, column=column, value=value)
        cell.font = Font(bold=True)
        cell.fill = fill


def _autosize_columns(sheet) -> None:
    for column_cells in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            max_length = max(max_length, len(str(cell.value or "")))
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 48)
